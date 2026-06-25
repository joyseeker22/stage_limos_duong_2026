import pandas as pd
import numpy as np
import os    #Operating system interface
from openpyxl import load_workbook


def read_excel_file(filename):
    
    print(f"\Reading file: {filename}")  
    print("-" * 50)
    
    info_sheet = load_workbook(filename).active
    print(f"Sheet 1 read: Contains dimensions")
    num_jobs = info_sheet.cell(row=1, column=2).value
    num_machines = info_sheet.cell(row=2, column=2).value
    print(f"  - Number of jobs: {num_jobs}")
    print(f"  - Number of machines: {num_machines}")
    num_jobs = int(num_jobs)
    num_machines = int(num_machines)
    
    processing_df = pd.read_excel(filename, sheet_name=1, index_col=0)
    print(f"Sheet 2 read: Processing Times ({processing_df.shape[0]} jobs × {processing_df.shape[1]} operations)")  #processing_df.shape: Returns (rows, columns) as a tuple
    
    processing_times = {}
    for job_name in processing_df.index:
        processing_times[str(job_name)] = processing_df.loc[job_name].tolist()
    
    machine_df = pd.read_excel(filename, sheet_name=2, index_col=0)
    print(f"Sheet 3 read: Machine Sequences ({machine_df.shape[0]} jobs × {machine_df.shape[1]} operations)")
    
    machine_sequences = {}
    for job_name in machine_df.index:
        machine_sequences[str(job_name)] = machine_df.loc[job_name].tolist()
    
    due_df = pd.read_excel(filename, sheet_name=3, index_col=0)
    print(f"Sheet 4 read: Due Dates ({len(due_df)} jobs)")
    
    due_dates = {}
    for job_name in due_df.index:
        job_str = str(job_name)
        priority = int(due_df.loc[job_name].iloc[0])
        due_date = int(due_df.loc[job_name].iloc[1])
        
        due_dates[job_str] = {
            'priority': priority,
            'due_date': due_date
        }
    
    return {
        'num_jobs': num_jobs,
        'num_machines': num_machines,
        'processing_times': processing_times,
        'machine_sequences': machine_sequences,
        'due_dates': due_dates
    }


def create_2factory_dataset(original_data):
    
    print(f"\Creating 2-factory dataset")
    print("-" * 50)
    
    num_machines = original_data['num_machines']
    processing_times = original_data['processing_times']
    machine_sequences = original_data['machine_sequences']
    due_dates = original_data['due_dates']
    
    # Generate random distance between factories (1-10 km)
    distance_km = np.random.randint(1, 11)
    print(f" Distance between factories: {distance_km} km")
    
    machine_offset = num_machines
    all_rows = []
    
    for job_name in processing_times.keys():
        print(f"Processing {job_name}...")
        
        job_times = processing_times[job_name]
        job_machines = machine_sequences.get(job_name, list(range(1, len(job_times)+1)))
        
        job_info = due_dates.get(job_name, {})
        job_priority = job_info.get('priority', 1)
        job_due_date = job_info.get('due_date', 999)
        
        for op_index in range(len(job_times)):
            processing_time = job_times[op_index]
            machine_type = int(job_machines[op_index]) if op_index < len(job_machines) else (op_index % num_machines + 1)
            
            row = {
                'Job': job_name,
                'Operation_Index': op_index + 1,  # Start from 1
                'Machine_Type': machine_type,
                'Factory1_Machine': machine_type,
                'Factory2_Machine': machine_type + machine_offset,
                'Processing_Time': processing_time,
                'Priority': job_priority,
                'Due_Date': job_due_date,
                'Distance_km': round(distance_km, 1),
            }
            
            all_rows.append(row)
    
    df_result = pd.DataFrame(all_rows)
    
    column_order = [
        'Job', 
        'Operation_Index',
        'Machine_Type',
        'Factory1_Machine', 
        'Factory2_Machine',
        'Processing_Time',
        'Priority',
        'Due_Date',
        'Distance_km'
    ]
    df_result = df_result[column_order]
    
    print(f"\ Created dataset with {len(df_result)} rows")
    
    return df_result, distance_km


def save_results(df_main, original_data, distance_km, output_filename):
    
    print(f"\ Saving results...")
    print("-" * 50)

    print(f"\ Saving Excel: {output_filename}")
    
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:

        df_main.to_excel(writer, sheet_name='2Factory_Dataset', index=False)
        print("2Factory_Dataset")
        
        info_data = {
            'Parameter': [
                'Number of Jobs',
                'Number of Machines per Factory',
                'Distance between Factories (km)',
                'Total Operations'
            ],
            'Value': [
                original_data['num_jobs'],
                original_data['num_machines'],
                distance_km,
                len(df_main)
            ]
        }
        info_df = pd.DataFrame(info_data)
        info_df.to_excel(writer, sheet_name='Factory_Info', index=False)
        print(" Factory_Info")
    
    print(f"Excel saved: {output_filename}")

    
def main():
    
    # STEP 1: Ask for file path
    print("\Please enter the path to your Excel file.")
    print("-" * 70)
    
    while True:
        INPUT_EXCEL = input("File path: ").strip()
        INPUT_EXCEL = INPUT_EXCEL.strip('"').strip("'")
        
        if os.path.exists(INPUT_EXCEL):
            print(f"✅ File found: {INPUT_EXCEL}")
            break
        else:
            print(f"File not found: {INPUT_EXCEL}")
            print("   Please check the path and try again.")
            print()
    
    # STEP 2: Output file name
    print(f"\Output file name (press Enter for default):")
    default_output = "output_2factory_data.xlsx"
    OUTPUT_EXCEL = input(f"   Output name [{default_output}]: ").strip()
    
    if not OUTPUT_EXCEL:
        OUTPUT_EXCEL = default_output
        print(f"   Using default: {OUTPUT_EXCEL}")
    
    if not OUTPUT_EXCEL.endswith('.xlsx'):
        OUTPUT_EXCEL += '.xlsx'
    
    # STEP 3: Confirmation
    print("\n" + "=" * 70)
    print("📋 SUMMARY")
    print("=" * 70)
    print(f"   Input file:    {INPUT_EXCEL}")
    print(f"   Output file:   {OUTPUT_EXCEL}")
    print("=" * 70)
    
    proceed = input("\n✅ Proceed? (Y/n): ").strip().lower()
    if proceed and proceed != 'y':
        print("❌ Cancelled by user.")
        return
    
    # STEP 4: Read Excel
    print("\n" + "=" * 70)
    print("STEP 1: Reading your Excel file")
    print("=" * 70)
    
    original_data = read_excel_file(INPUT_EXCEL)
    
    # STEP 5: Create 2-factory dataset
    print("\n" + "=" * 70)
    print("STEP 2: Creating 2-factory dataset")
    print("=" * 70)
    
    df_main, distance = create_2factory_dataset(original_data)
    
    # STEP 6: Save
    print("\n" + "=" * 70)
    print("STEP 3: Saving results")
    print("=" * 70)
    
    save_results(df_main, original_data, distance, OUTPUT_EXCEL)

if __name__ == "__main__":
    main()
